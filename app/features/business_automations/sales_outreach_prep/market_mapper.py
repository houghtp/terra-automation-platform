import os
import re
import pandas as pd
import plotly.express as px
import random
import requests
from openpyxl import Workbook
import tempfile
from datetime import datetime
from io import BytesIO
from typing import List, Dict
from urllib.parse import quote_plus
import sys
import json

"""
Market Mapper: AI-powered competitive intelligence tool

This tool uses OpenAI, web search, and data scraping to generate:
- Competitive landscape analysis
- Executive contact information
- Market positioning quadrants

All API functionality is centralized in TerraITAgent/utils/api_clients.py for consistency.
"""

# Add TerraITAgent utils directory to the path to import utilities
sys.path.append(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "TerraITAgent"))
from utils.content_scrapers import search_firecrawl
from utils.exporters import export_to_excel as terra_export_to_excel
from utils.api_clients import query_openai, find_email, get_openai_client, load_api_keys

# --- CONFIGURATION ---
# Load all API keys from environment using centralized utility
api_keys = load_api_keys()
REPORTS_DIR = "reports"
os.makedirs(REPORTS_DIR, exist_ok=True)

# --- GPT QUERY ---
def query_openai_for_market_data(company_name: str) -> tuple[pd.DataFrame, str, list[str]]:
    prompt = f"""
You are a senior market analyst.

Part 1: Provide a JSON list of 5–10 competitors (direct or indirect) to {company_name}.
Each object should include:
- Company name
- Short description (max 25 words)
- Estimated market size: "Low", "Medium", or "High"
- Product breadth: "Niche", "Moderate", or "Broad"
- Innovation score: 1–10
- Completeness of vision: 1–10
- Ability to execute: 1–10

Part 2: Provide an executive summary (150 words max) of the market landscape, labeled "Executive Summary:"
Part 3: Provide 3–5 bullet-point insights labeled "AI Insights:".
Each insight should highlight non-obvious patterns, outliers, disruptors, or strategic risks. Avoid generic observations.
"""

    content = query_openai(prompt=prompt, model="gpt-4", temperature=0.7)

    try:
        parts = content.split("Executive Summary:")
        json_str = parts[0].strip()
        summary_split = parts[1].split("AI Insights:")
        summary = summary_split[0].strip()
        insights = [line.strip('- ').strip() for line in summary_split[1].strip().split('\n') if line.strip()]

        import json
        json_data = json.loads(json_str[json_str.find("["):json_str.find("]") + 1])
        df = pd.DataFrame(json_data)

        df.columns = [col.strip().lower() for col in df.columns]
        df.rename(columns={
            "company name": "Company name",
            "short description": "Short description",
            "estimated market size": "Estimated market size",
            "product breadth": "Product Breadth",
            "innovation score": "Innovation Score",
            "completeness of vision": "Completeness of Vision",
            "ability to execute": "Ability to Execute"
        }, inplace=True)

        # Append the target company to the list for comparison
        df = pd.concat([
            df,
            pd.DataFrame([{
                "Company name": company_name,
                "Short description": "Subject of analysis",
                "Estimated market size": "High",
                "Product Breadth": "Broad",
                "Innovation Score": 7,
                "Completeness of Vision": 8,
                "Ability to Execute": 8
            }])
        ], ignore_index=True)

        return df, summary, insights

    except Exception as e:
        raise ValueError(f"Error parsing GPT response: {e}\n\nRaw Content:\n{content}")

def jitter_duplicate_points(df: pd.DataFrame, x_col: str, y_col: str) -> pd.DataFrame:
    df[x_col] = df[x_col].astype(float)
    df[y_col] = df[y_col].astype(float)
    seen = {}
    for idx, row in df.iterrows():
        key = (row[x_col], row[y_col])
        if key in seen:
            df.at[idx, x_col] += round(random.uniform(-0.3, 0.3), 2)
            df.at[idx, y_col] += round(random.uniform(-0.3, 0.3), 2)
        else:
            seen[key] = True
    return df

def generate_quadrant_chart(df: pd.DataFrame, company_name: str) -> str:
    df = jitter_duplicate_points(df, "Completeness of Vision", "Ability to Execute")

    fig = px.scatter(
        df,
        x="Completeness of Vision",
        y="Ability to Execute",
        text="Company name",
        color="Estimated market size",
        hover_data=["Short description", "Product Breadth", "Innovation Score"],
        title=f"Gartner-Style Market Quadrant: {company_name}",
    )

    fig.update_traces(
        textposition="top center",
        textfont=dict(size=10),
        marker=dict(size=12, line=dict(width=0.5, color="DarkSlateGrey"))
    )

    fig.add_shape(type="line", x0=5, x1=5, y0=0, y1=10, line=dict(dash="dash", color="gray"))
    fig.add_shape(type="line", x0=0, x1=10, y0=5, y1=5, line=dict(dash="dash", color="gray"))

    fig.update_layout(
        xaxis_title="Completeness of Vision",
        yaxis_title="Ability to Execute",
        xaxis=dict(range=[0, 10]),
        yaxis=dict(range=[0, 10]),
        margin=dict(l=40, r=40, t=60, b=40),
        height=600,
        width=900
    )

    chart_path = os.path.join(tempfile.gettempdir(), "gartner_quadrant_chart.png")
    fig.write_image(chart_path)
    assert os.path.exists(chart_path), f"Failed to export quadrant chart to {chart_path}"
    return chart_path

def fetch_logo_image(company_name: str) -> BytesIO | None:
    """Fetch company logo from clearbit API
    
    Args:
        company_name: Name of company to fetch logo for
        
    Returns:
        BytesIO object containing the logo image data or None if not found
    """
    try:
        domain = company_name.lower().replace(" ", "") + ".com"
        logo_url = f"https://logo.clearbit.com/{domain}"
        response = requests.get(logo_url, timeout=5)
        if response.status_code == 200:
            return BytesIO(response.content)
    except Exception:
        pass
    return None
    """Extract job title from a snippet of text
    
    Args:
        snippet: Text snippet to search for job title
        
    Returns:
        Extracted job title or empty string if not found
    """
    patterns = [
        r"\\bI am a (.+?) with\\b",
        r"\\bI am an (.+?) with\\b",
        r"\\bI work as (.+?) at\\b",
        r"\\bserving as (.+?) at\\b",
        r"\\bcurrently a[n]* (.+?) at\\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, snippet, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return ""

def extract_title_from_snippet(snippet: str) -> str:
    patterns = [
        r"\\bI am a (.+?) with\\b",
        r"\\bI am an (.+?) with\\b",
        r"\\bI work as (.+?) at\\b",
        r"\\bserving as (.+?) at\\b",
        r"\\bcurrently a[n]* (.+?) at\\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, snippet, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return ""

def find_executives_via_openai_websearch(company: str, country: str, role_query: str) -> List[Dict[str, str]]:
    """
    Finds LinkedIn executives using OpenAI's web_search tool (Responses API).
    This does not consume SerpAPI credits.
    """
    query = f'site:linkedin.com/in "{company}" {country} ({role_query})'
    client = get_openai_client()

    try:
        response = client.responses.create(
            model="gpt-4o",
            input=f"Using the query: {query}, list up to 10 LinkedIn profiles as JSON with keys: name, title, region, and source (the profile URL).",
            tools=[{"type": "web_search"}]
        )

        tool_outputs = response.choices[0].message.tool_calls
        if not tool_outputs:
            print("[✗] No tool calls returned.")
            return []

        # Extract and parse result
        raw_text = response.choices[0].message.content.strip()

        import json
        try:
            data = json.loads(raw_text)
            if isinstance(data, list):
                return data
            else:
                print("[!] Response wasn't a list, returning empty.")
                return []
        except Exception as e:
            print(f"[!] Could not parse JSON output: {e}")
            print("Raw response:")
            print(raw_text)
            return []

    except Exception as e:
        print(f"[!] OpenAI web search failed: {e}")
        return []

def find_executives_via_firecrawl(company: str, country: str, role_query: str, max_results: int = 10) -> list[dict]:
    """
    Find executive LinkedIn profiles using Firecrawl search API
    
    Args:
        company: Company name to search for
        country: Country or region to filter by
        role_query: Role or position query (e.g., "CEO OR CTO")
        max_results: Maximum number of results to return
        
    Returns:
        List of dictionaries with executive information
    """
    firecrawl_api_key = api_keys.get("firecrawl")
    if not firecrawl_api_key:
        print("[!] Missing Firecrawl API key.")
        return []
    
    # Construct the search query for LinkedIn profiles
    query = f'site:linkedin.com/in "{company}" {country} ({role_query})'
    print(f"[•] Searching Firecrawl with query: {query}")
    
    try:        # Use search_firecrawl directly from TerraITAgent
        search_result = search_firecrawl(query, api_key=api_keys.get("firecrawl"), limit=max_results)
        
        execs = []
        # Process the search results
        if hasattr(search_result, 'data'):
            for result in search_result.data:
                title_text = getattr(result, 'title', '')
                description = getattr(result, 'description', '')
                link = getattr(result, 'url', '')
                
                if "linkedin.com/in" in link:
                    # Extract name and title from the title text (format usually: "Name - Title at Company")
                    parts = title_text.split(" - ")
                    name = parts[0].strip() if len(parts) >= 1 else ""
                    raw_title = parts[1].strip() if len(parts) >= 2 else ""
                    fallback_title = extract_title_from_snippet(description)
                    title = raw_title if raw_title else fallback_title
                    
                    execs.append({
                        "name": name,
                        "title": title,
                        "region": country,
                        "source": link,
                        "snippet": description
                    })
        
        print(f"[✓] Found {len(execs)} executives via Firecrawl")
        return execs
        
    except Exception as e:
        print(f"[!] Firecrawl search error: {str(e)}")
        # Print traceback for debugging
        import traceback
        traceback.print_exc()
        return []

def get_company_domains_via_gpt(company_name: str) -> list[str]:
    """
    Get the official domain names for a company using GPT-4.
    
    Args:
        company_name: Name of the company
        
    Returns:
        List of verified domain names
    """
    print(f"[•] Looking up domains for {company_name}...")
    
    # More focused prompt with explicit constraints
    prompt = f"""
You are a corporate domain name expert with access to factual company information.

For the company "{company_name}", list ONLY the official website domains the company actively uses. 
Focus ONLY on these domain types:
1. Primary corporate domain (e.g., microsoft.com)
2. Consumer brand domains (e.g., xbox.com)
3. Major division domains (e.g., azure.microsoft.com)
4. Common short aliases (e.g., ms.com)

IMPORTANT CONSTRAINTS:
- Return between 1-5 domains maximum
- Only include domains you're highly confident actually exist
- Do NOT include subdomains (e.g., investors.company.com)
- Do NOT include social media, third-party, or partner domains
- Do NOT hallucinate domains - if unsure, include fewer domains rather than guessing

Format your response EXACTLY like this without ANY additional text:
- domain1.com
- domain2.com
"""

    try:
        # Call GPT-4 using the centralized utility
        content = query_openai(
            prompt=prompt,
            model="gpt-4",
            temperature=0.1,
            system_message="You provide factual domain information only."
        )
        
        # Extract domains from the response
        domains = [
            line.strip("- ").strip()
            for line in content.split("\n")
            if line.strip() and "." in line
        ]
        
        # Filter out obviously invalid domains (extra verification)
        valid_domains = []
        for domain in domains[:5]:  # Hard limit to 5 domains
            # Basic pattern check
            if (
                len(domain) > 3 and                                       # At least 4 chars
                "." in domain and                                         # Has a dot
                not domain.startswith(".") and                            # Doesn't start with a dot
                not domain.endswith(".") and                              # Doesn't end with a dot
                len(domain.split(".")[0]) >= 2 and                        # At least 2 chars before dot
                not any(s in domain for s in [".....", "----", "xxxxx"])  # No obvious placeholders
            ):
                valid_domains.append(domain.lower())
        
        if valid_domains:
            print(f"[✓] Found {len(valid_domains)} domains: {', '.join(valid_domains)}")
        else:
            print(f"[!] No valid domains found for {company_name}")
            
        return valid_domains

    except Exception as e:
        print(f"[!] Failed to fetch domains via GPT: {e}")
        return []
    
def clean_full_name(name: str) -> str:
    # Remove suffixes like ", CFA", ", PhD", "CFA", "MBA"
    name = re.sub(r",.*$", "", name)
    credentials = ["CFA", "PhD", "MBA", "MD", "Esq", "CPA"]
    for cred in credentials:
        name = re.sub(rf"\b{cred}\b", "", name, flags=re.IGNORECASE)
    return name.strip()

def get_execs_for_all_companies(df: pd.DataFrame, countries: list[str], roles: str) -> dict[str, list[dict]]:
    execs_by_company = {}
    main_company = df.iloc[-1]["Company name"]
    roles_set = set([r.lower().strip() for r in roles.split("OR")])

    for _, row in df.iterrows():
        company = row["Company name"]
        is_main_company = company == main_company
        domains = get_company_domains_via_gpt(company)
        execs_by_company[company] = []

        if not is_main_company:

            for country in countries:
                # Construct search query for LinkedIn profiles
                query = f'site:linkedin.com/in "{company}" {country} ({roles})'
                print(f"[•] Searching for {company} executives in {country}...")
                  # Use the enhanced search_firecrawl function directly with profile extraction
                search_result = search_firecrawl(
                    query=query,
                    api_key=api_keys.get("firecrawl"),
                    limit=20,
                    extract_profiles=True
                )
                
                if not search_result.get("success", False):
                    print(f"[!] Firecrawl search failed: {search_result.get('error', 'Unknown error')}")
                    continue
                    
                # Process the search results
                for result in search_result.get("results", []):
                    if "profile" not in result:
                        continue
                        
                    profile = result["profile"]
                    full_name = profile["name"]
                    cleaned_name = clean_full_name(full_name)
                    raw_title = profile["title"]
                    snippet = profile.get("snippet", "")
                    
                    # Check if the role matches
                    exec_title_lower = raw_title.lower()
                    #matches_role = any(role in exec_title_lower for role in roles_set)
                    
                    # if not matches_role:
                    #     continue
                        
                    email = "N/A"
                    matched_domain = "N/A"

                    # Try email lookup only if not the main company
                    if not is_main_company:
                        for domain in domains:
                            email = find_email(cleaned_name, domain)
                            if email != "N/A":
                                matched_domain = domain
                                break

                    execs_by_company[company].append({
                        "name": full_name,
                        "title": raw_title,
                        "region": country,
                        "domain": matched_domain,
                        "email": email,
                        "source": profile["source"],
                        "snippet": snippet
                    })
                
                print(f"[✓] Found {len(execs_by_company[company])} executives for {company} in {country}")

    return execs_by_company

def export_to_excel(company_name: str, df: pd.DataFrame, summary: str, insights: list[str], execs_by_company: dict):
    """
    Export market data to Excel using the TerraITAgent shared utility function
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"market_report_{company_name.lower()}_{timestamp}.xlsx"
    
    # Create executive data in a format that's easy to process
    exec_data = []
    for company, execs in execs_by_company.items():
        for exec in execs:
            exec_data.append({
                "Company": company,
                "Name": exec.get("name"),
                "Title": exec.get("title"),
                "Region": exec.get("region"),
                "Email": exec.get("email"),
                "Source": exec.get("source"),
                "Snippet": exec.get("snippet")
            })
    
    # Format executive summary data
    summary_data = [
        ["Executive Summary"],
        [summary],
        [],
        ["AI Insights"]
    ]
    for bullet in insights:
        summary_data.append([f"- {bullet}"])
    
    # Create data structure for the shared export utility
    sheets_data = {
        "Executive Summary": summary_data,
        "Market Data": df,  # The utility can handle DataFrame objects directly
        "Executives": exec_data  # List of dictionaries will be formatted properly
    }
    
    try:
        # Use the shared utility to export the data
        file_path = terra_export_to_excel(sheets_data, filename, REPORTS_DIR)
        print(f"[✔] Excel report saved: {file_path}")
        return file_path
    except Exception as e:
        print(f"[!] Error using shared export utility: {e}")
        print("[*] Falling back to direct implementation...")
        
        # Fallback to original implementation if needed
        local_file_path = os.path.join(REPORTS_DIR, filename)
        wb = Workbook()

        # Executive Summary
        summary_ws = wb.active
        summary_ws.title = "Executive Summary"
        summary_ws.append(["Executive Summary"])
        summary_ws.append([summary])
        summary_ws.append([])
        summary_ws.append(["AI Insights"])
        for bullet in insights:
            summary_ws.append([f"- {bullet}"])

        # Market Data
        market_ws = wb.create_sheet(title="Market Data")
        market_ws.append(df.columns.tolist())
        for _, row in df.iterrows():
            market_ws.append(row.tolist())

        # Executives
        exec_ws = wb.create_sheet(title="Executives")
        exec_ws.append(["Company", "Name", "Title", "Region", "Email", "Source","Snippet"])
        for company, execs in execs_by_company.items():
            for exec in execs:
                exec_ws.append([
                    company,
                    exec.get("name"),
                    exec.get("title"),
                    exec.get("region"),
                    exec.get("email"),
                    exec.get("source"),
                    exec.get("snippet")
                ])

        wb.save(local_file_path)
        print(f"[✔] Excel report saved using fallback method: {local_file_path}")
        return local_file_path

# --- MAIN ---
def main():
    print("=== AI Market Mapper: Excel Export Edition ===")
    company_name = input("Enter a company name: ").strip()
    raw_roles = input("Enter roles to search for (comma-separated): ").strip()

     # Convert to "VP OR Director OR Chief" format
    roles = " OR ".join(role.strip() for role in raw_roles.split(",") if role.strip())

    print(f"\n[•] Querying GPT-4 for market landscape around '{company_name}'...")
    df, summary, insights = query_openai_for_market_data(company_name)

    # print("[•] Generating quadrant chart...")
    # chart_path = generate_quadrant_chart(df, company_name)

    print("[•] Finding executives by region for each company...")
    countries = ["Boston"]
    execs_by_company = get_execs_for_all_companies(df, countries, roles)

    print("[•] Creating Excel report...")
    export_to_excel(company_name, df, summary, insights, execs_by_company)

    print("\n[✔] Done.")

if __name__ == "__main__":
    main()
